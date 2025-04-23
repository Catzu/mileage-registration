import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
import os
import openrouteservice
import requests
import time
import json
from datetime import datetime
import threading
import re

class DeliveryDistanceCalculator:
    # Configuration constants
    DEFAULT_CONFIG = {
        "ors_api_key": "#",
        "cafetaria_address": "#",
        "excel_file": "excel.xlsx",
        "rate_per_km": 0.23,
        "max_distance_warn": 30
    }
    CONFIG_FILE = "config.json"
    
    # Excel header definition
    EXCEL_HEADER = [
        'Datum',
        'Postcode cafetaria',
        'Postcode klant',
        'Adres klant (indien postcode vergeten te vragen)',
        'Aantal kilom heen en terug',
        'Vergoeding per km',
        'Totaalbedrag (niet invullen is formule)',
        'Fout melding'
    ]

    def __init__(self, root):
        self.root = root
        self.root.title("Leveringsafstand Berekening")
        self.root.geometry("600x650")
        
        # Load configuration
        self.config = self.load_config()
        
        # Initialize client only when needed
        self.ors_client = None
        
        # Initialize cafetaria coordinates cache
        self.cafetaria_coords = None
        
        # Setup UI
        self.create_widgets()
        
        # Setup Excel file
        self.ensure_excel_file_exists()

    def load_config(self):
        """Load configuration from file or create default"""
        if os.path.exists(self.CONFIG_FILE):
            try:
                with open(self.CONFIG_FILE, 'r') as f:
                    return json.load(f)
            except:
                messagebox.showwarning("Config Error", "Configuratiebestand is beschadigd. Standaardwaarden worden gebruikt.")
        
        # Save default config
        with open(self.CONFIG_FILE, 'w') as f:
            json.dump(self.DEFAULT_CONFIG, f, indent=4)
        
        return self.DEFAULT_CONFIG.copy()

    def save_config(self):
        """Save current configuration to file"""
        with open(self.CONFIG_FILE, 'w') as f:
            json.dump(self.config, f, indent=4)

    def ensure_excel_file_exists(self):
        """Setup Excel file if it doesn't exist"""
        if not os.path.exists(self.config["excel_file"]):
            wb = Workbook()
            ws = wb.active
            ws.title = "Kilometerregistratie"
            ws.append(self.EXCEL_HEADER)
            wb.save(self.config["excel_file"])

    def create_widgets(self):
        """Create all UI elements"""
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Create main tab
        self.main_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.main_tab, text="Leveringen")
        
        # Create settings tab
        self.settings_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_tab, text="Instellingen")
        
        # Setup main tab
        self.setup_main_tab()
        
        # Setup settings tab
        self.setup_settings_tab()
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Gereed")
        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, relief="sunken", anchor="w")
        self.status_bar.pack(side="bottom", fill="x")

    def setup_main_tab(self):
        """Setup the main tab with delivery inputs"""
        # Frame for date
        date_frame = ttk.Frame(self.main_tab)
        date_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(date_frame, text="Datum (dd-mm-jjjj):").pack(side="left")
        self.date_entry = ttk.Entry(date_frame)
        self.date_entry.pack(side="left", padx=5)
        # Set default date to today
        today = datetime.now().strftime("%d-%m-%Y")
        self.date_entry.insert(0, today)
        
        # Button to select Excel file
        file_frame = ttk.Frame(self.main_tab)
        file_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(file_frame, text="Excel bestand:").pack(side="left")
        self.file_var = tk.StringVar(value=self.config["excel_file"])
        ttk.Entry(file_frame, textvariable=self.file_var, width=30).pack(side="left", padx=5)
        ttk.Button(file_frame, text="Blader...", command=self.select_excel_file).pack(side="left")
        
        # Frame for delivery count
        count_frame = ttk.Frame(self.main_tab)
        count_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(count_frame, text="Aantal leveringen:").pack(side="left")
        self.deliveries_entry = ttk.Entry(count_frame, width=5)
        self.deliveries_entry.pack(side="left", padx=5)
        self.deliveries_entry.insert(0, "1")
        ttk.Button(count_frame, text="Genereer Leveringsvelden", 
                   command=self.generate_fields).pack(side="left", padx=5)
        
        # Scrollable frame for delivery entries
        self.delivery_container = ttk.Frame(self.main_tab)
        self.delivery_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Canvas for scrolling
        self.canvas = tk.Canvas(self.delivery_container)
        scrollbar = ttk.Scrollbar(self.delivery_container, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        
        # Frame within canvas for delivery fields
        self.delivery_frame = ttk.Frame(self.canvas)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.delivery_frame, anchor="nw")
        
        self.canvas.bind('<Configure>', self.on_canvas_configure)
        self.delivery_frame.bind('<Configure>', self.on_frame_configure)
        
        # Save button
        ttk.Button(self.main_tab, text="Opslaan in Excel", 
                   command=self.save_to_excel_threaded).pack(pady=10)
        
        # Initialize entries list
        self.postcode_entries = []
        self.address_entries = []
        
        # Generate the initial field
        self.generate_fields()

    def setup_settings_tab(self):
        """Setup the settings tab"""
        settings_frame = ttk.Frame(self.settings_tab)
        settings_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # API Key
        ttk.Label(settings_frame, text="OpenRouteService API Key:").grid(row=0, column=0, sticky="w", pady=5)
        self.api_key_var = tk.StringVar(value=self.config["ors_api_key"])
        ttk.Entry(settings_frame, textvariable=self.api_key_var, width=40).grid(row=0, column=1, sticky="w", padx=5)
        
        # Cafetaria Address
        ttk.Label(settings_frame, text="Cafetaria adres:").grid(row=1, column=0, sticky="w", pady=5)
        self.cafetaria_var = tk.StringVar(value=self.config["cafetaria_address"])
        ttk.Entry(settings_frame, textvariable=self.cafetaria_var, width=40).grid(row=1, column=1, sticky="w", padx=5)
        
        # Rate per km
        ttk.Label(settings_frame, text="Vergoeding per km (€):").grid(row=2, column=0, sticky="w", pady=5)
        self.rate_var = tk.StringVar(value=str(self.config["rate_per_km"]))
        ttk.Entry(settings_frame, textvariable=self.rate_var, width=10).grid(row=2, column=1, sticky="w", padx=5)
        
        # Max distance warning
        ttk.Label(settings_frame, text="Max afstand voor waarschuwing (km):").grid(row=3, column=0, sticky="w", pady=5)
        self.max_distance_var = tk.StringVar(value=str(self.config["max_distance_warn"]))
        ttk.Entry(settings_frame, textvariable=self.max_distance_var, width=10).grid(row=3, column=1, sticky="w", padx=5)
        
        # Save settings button
        ttk.Button(settings_frame, text="Instellingen Opslaan", 
                   command=self.save_settings).grid(row=4, column=0, columnspan=2, pady=20)
        
        # Test connection button
        ttk.Button(settings_frame, text="Test API Verbinding", 
                   command=self.test_ors_connection).grid(row=5, column=0, columnspan=2)

    def on_canvas_configure(self, event):
        """Handle canvas resize"""
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def on_frame_configure(self, event):
        """Update the scrollregion to encompass the inner frame"""
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def select_excel_file(self):
        """Let user select Excel file"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=self.config["excel_file"]
        )
        if file_path:
            self.file_var.set(file_path)
            self.config["excel_file"] = file_path
            self.save_config()
            self.ensure_excel_file_exists()

    def save_settings(self):
        """Save settings to config file"""
        try:
            self.config["ors_api_key"] = self.api_key_var.get()
            self.config["cafetaria_address"] = self.cafetaria_var.get()
            self.config["rate_per_km"] = float(self.rate_var.get().replace(',', '.'))
            self.config["max_distance_warn"] = float(self.max_distance_var.get().replace(',', '.'))
            self.config["excel_file"] = self.file_var.get()
            
            # Reset client so it uses the new API key
            self.ors_client = None
            # Reset cafetaria coords cache
            self.cafetaria_coords = None
            
            self.save_config()
            messagebox.showinfo("Instellingen", "Instellingen opgeslagen!")
        except ValueError:
            messagebox.showerror("Invoerfout", "Voer geldige getallen in voor vergoeding en max afstand.")

    def test_ors_connection(self):
        """Test the OpenRouteService API connection"""
        self.status_var.set("API verbinding testen...")
        self.root.update()
        
        # Initialize client with current API key
        api_key = self.api_key_var.get()
        client = openrouteservice.Client(key=api_key)
        
        try:
            # Simple request to test connection
            client.directions(
                coordinates=[(4.8, 52.3), (4.9, 52.4)],
                profile='driving-car',
                format='geojson'
            )
            messagebox.showinfo("API Test", "Verbinding succesvol!")
            self.status_var.set("API verbinding getest: OK")
        except Exception as e:
            messagebox.showerror("API Test", f"Verbindingsfout: {str(e)}")
            self.status_var.set("API verbinding getest: FOUT")

    def generate_fields(self):
        """Create input fields dynamically based on number of deliveries"""
        # Clear the frame
        for widget in self.delivery_frame.winfo_children():
            widget.destroy()
        
        try:
            num = int(self.deliveries_entry.get())
            if num < 1:
                raise ValueError("Aantal moet minimaal 1 zijn")
        except ValueError:
            messagebox.showerror("Invoerfout", "Aantal leveringen moet een positief getal zijn.")
            self.deliveries_entry.delete(0, tk.END)
            self.deliveries_entry.insert(0, "1")
            num = 1
        
        # Header row
        ttk.Label(self.delivery_frame, text="Postcode", width=10).grid(row=0, column=0, padx=5, pady=2)
        ttk.Label(self.delivery_frame, text="Alternatief adres", width=30).grid(row=0, column=1, padx=5, pady=2)
        
        self.postcode_entries = []
        self.address_entries = []
        
        # Create rows
        for i in range(num):
            # Dutch postal code is formatted as "1234 AB"
            postcode_label = ttk.Label(self.delivery_frame, text=f"#{i+1}:")
            postcode_label.grid(row=i+1, column=0, sticky="e", padx=5, pady=5)
            
            postcode_frame = ttk.Frame(self.delivery_frame)
            postcode_frame.grid(row=i+1, column=0, padx=5, pady=2)
            
            postcode_entry = ttk.Entry(self.delivery_frame, width=10)
            postcode_entry.grid(row=i+1, column=0, padx=5, pady=2)
            self.postcode_entries.append(postcode_entry)
            
            address_entry = ttk.Entry(self.delivery_frame, width=40)
            address_entry.grid(row=i+1, column=1, padx=5, pady=2)
            self.address_entries.append(address_entry)

    def nominatim_geocode(self, address):
        """Geocode an address using Nominatim"""
        url = "https://nominatim.openstreetmap.org/search"
        params = {
            'q': address,
            'format': 'json',
            'countrycodes': 'nl',
            'limit': 1
        }
        try:
            response = requests.get(url, params=params, headers={'User-Agent': 'DistanceCalculatorApp'})
            data = response.json()
            if data:
                return float(data[0]['lon']), float(data[0]['lat'])
            else:
                return None
        except Exception as e:
            self.status_var.set(f"Geocoding fout: {str(e)}")
            return None

    def get_ors_client(self):
        """Get or initialize the ORS client"""
        if self.ors_client is None:
            self.ors_client = openrouteservice.Client(key=self.config["ors_api_key"])
        return self.ors_client

    def get_route_distance(self, from_coords, to_coords):
        """Calculate round-trip driving distance using ORS"""
        try:
            client = self.get_ors_client()
            route = client.directions(
                coordinates=[from_coords, to_coords],
                profile='driving-car',
                format='geojson'
            )
            # One-way distance in km × 2 for round-trip
            distance = route['features'][0]['properties']['segments'][0]['distance'] / 1000
            return round(distance * 2, 2)
        except Exception as e:
            self.status_var.set(f"Routefout: {str(e)}")
            return None

    def validate_postcode(self, postcode):
        """Validate Dutch postal code format (1234 AB)"""
        # Remove spaces and uppercase
        postcode = postcode.strip().upper()
        
        # Check if empty
        if not postcode:
            return False
            
        # Basic regex for Dutch postal code
        pattern = r'^\d{4}\s?[A-Z]{2}$'
        if re.match(pattern, postcode):
            # Format consistently as "1234 AB"
            if ' ' not in postcode:
                return postcode[:4] + ' ' + postcode[4:]
            return postcode
        return False

    def save_to_excel_threaded(self):
        """Start saving process in a separate thread to keep UI responsive"""
        threading.Thread(target=self.save_to_excel, daemon=True).start()

    def save_to_excel(self):
        """Fetch all inputs, calculate distances, and write data to Excel"""
        self.status_var.set("Bezig met opslaan...")
        self.root.update()
        
        # Validate date format
        date = self.date_entry.get()
        try:
            datetime.strptime(date, "%d-%m-%Y")
        except ValueError:
            messagebox.showerror("Datumfout", "Gebruik formaat: dd-mm-jjjj")
            self.status_var.set("Gereed")
            return
        
        # Get current Excel file path
        excel_file = self.file_var.get()
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Excel fout", f"Kon het Excel-bestand niet openen: {str(e)}")
            self.status_var.set("Gereed")
            return
        
        # Geocode cafetaria once and cache
        if self.cafetaria_coords is None:
            self.status_var.set("Cafetaria locatie ophalen...")
            self.root.update()
            
            self.cafetaria_coords = self.nominatim_geocode(self.config["cafetaria_address"])
            if not self.cafetaria_coords:
                messagebox.showerror("Locatiefout", "Kon cafetaria niet geoloceren.")
                self.status_var.set("Gereed")
                return
        
        from_coords = self.cafetaria_coords
        cafetaria_postcode = "XXXX XX" #Hard coded postcode
        
        # Process each delivery
        rows_added = 0
        for i, (postcode_entry, address_entry) in enumerate(zip(self.postcode_entries, self.address_entries)):
            klant_postcode = postcode_entry.get().strip()
            klant_address = address_entry.get().strip()
            
            # Update status
            self.status_var.set(f"Verwerken levering {i+1}...")
            self.root.update()
            
            # Skip empty entries
            if not klant_postcode and not klant_address:
                continue
                
            # Validate postcode if provided
            valid_postcode = None
            if klant_postcode:
                valid_postcode = self.validate_postcode(klant_postcode)
                if not valid_postcode:
                    if not klant_address:  # If no alternative address
                        messagebox.showwarning(
                            "Ongeldige postcode", 
                            f"Levering #{i+1}: '{klant_postcode}' is geen geldige Nederlandse postcode. Vul een alternatief adres in."
                        )
                        continue
            
            # Determine which address to use for geocoding
            search_address = None
            if valid_postcode:
                search_address = f"{valid_postcode}, Netherlands"
                klant_postcode = valid_postcode
            elif klant_address:
                search_address = f"{klant_address}, Netherlands"
            else:
                continue  # Skip if no valid address
                
            # Geocode the address
            to_coords = self.nominatim_geocode(search_address)
            if not to_coords:
                messagebox.showwarning(
                    "Adres niet gevonden", 
                    f"Levering #{i+1}: Kon adres '{search_address}' niet vinden."
                )
                continue
            
            # Respect Nominatim's usage policy
            time.sleep(1) 
            
            # Get route distance
            distance = self.get_route_distance(from_coords, to_coords)
            if distance is None:
                messagebox.showwarning(
                    "Afstandsfout", 
                    f"Levering #{i+1}: Kon afstand niet berekenen voor '{search_address}'."
                )
                continue
            
            # Check if the distance exceeds the warning threshold
            error_message = ""
            if distance > self.config["max_distance_warn"]:
                error_message = "Handmatig controleren"
            
            # Add row to Excel
            next_row = ws.max_row + 1
            formula = f"=E{next_row}*F{next_row}"  # Total = Distance × Rate
            ws.append([
                date, 
                cafetaria_postcode, 
                klant_postcode, 
                klant_address, 
                distance, 
                self.config["rate_per_km"], 
                formula, 
                error_message
            ])
            
            rows_added += 1
        
        # Save Excel file
        try:
            wb.save(excel_file)
            if rows_added > 0:
                messagebox.showinfo("Succes", f"{rows_added} leveringen opgeslagen in Excel!")
            else:
                messagebox.showinfo("Info", "Geen leveringen om op te slaan.")
        except Exception as e:
            messagebox.showerror("Excel fout", f"Kon niet opslaan naar Excel: {str(e)}")
        
        self.status_var.set("Gereed")

# Run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = DeliveryDistanceCalculator(root)
    root.mainloop()